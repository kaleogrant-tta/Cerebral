"""
discount_group_audit_report.py -- a local worklist of accounts holding more
than one discount entitlement.

    python discount_group_audit_report.py

Writes Discount_Group_Audit_<date>.xlsx next to this script.

WHY THIS IS A LOCAL FILE AND NOT A DASHBOARD TAB
------------------------------------------------
cerebral_dash.duckdb is uploaded to Streamlit Community Cloud and is reachable
by URL. publish.py deliberately strips every customer identifier before it
leaves this machine, and its PII guard refuses to publish an unrecognised text
column for exactly that reason.

Policing discount stacking needs names. So this report stays here: it reads
the audit export and the local ETL database, writes an .xlsx, and uploads
nothing. Treat the output as an internal HR/ops document -- it names customers
and the staff member who granted each entitlement.

WHAT IT FLAGS
-------------
  Sheet "Multi-group accounts"   every customer with 2+ open entitlements,
                                 with the groups, their home store, discounted
                                 basket count and discount taken.
  Sheet "Granted + granted"      the subset holding two or more GRANTED groups
                                 (neighbour business, staff, employee, first
                                 responder). These are the ones that should
                                 not exist -- a paid Frequent Flyer membership
                                 alongside a granted rate is a policy
                                 question; two granted rates is an error.
  Sheet "Concentrated use"       groups with <= 3 members but 20+ discounted
                                 baskets: a rate meant for a business's staff
                                 being run as one person's standing discount.
  Sheet "Combinations"           how many accounts hold each pairing.
  Sheet "Notes"                  limits of the data, in the file itself.

MEMBERSHIP MODEL
----------------
Open memberships only: the last recorded action for that customer x group is
an Added. A customer removed from a group is not flagged.

Frequent Flyer is a PURCHASED membership ($100), not an earned tier, so it is
labelled "Paid membership" and counted separately from granted entitlements
rather than being treated as a non-issue.
"""

from __future__ import annotations

import datetime
import glob
import os
import sys
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = Path(os.path.expanduser("~/cerebral/tta.duckdb"))
HEADER_ROW = 3
ARIAL = "Arial"

STORE_LABEL = {1: "DTBK", 2: "Fifth Avenue", 3: "SoHo", 4: "Union Square"}

PAID = {"travel club frequent flyer", "travel club", "frequent flyer"}


def kind(name: str) -> str:
    low = str(name or "").strip().lower()
    if low in PAID:
        return "Paid membership"
    if "employee" in low:
        return "Employee"
    if "first responder" in low or "veteran" in low:
        return "First responder / veteran"
    if "retail worker" in low or "friends and family" in low:
        return "Staff / friends & family"
    if ("drinks on us" in low or "drinksonus" in low
            or "drink on us" in low):
        return "Neighbour business"
    if low.startswith(("soho -", "soho-", "5th ave", "5thave", "usq ",
                       "dtbk ", "fifth ave")):
        return "Neighbour business"
    return "Other"


def find_audit() -> Path | None:
    pats = ["**/*Discount Group Audit*.xls*", "**/*Discount_Group_Audit*.xls*"]
    for root in (".", os.path.expanduser("~/cerebral")):
        for pat in pats:
            hits = glob.glob(os.path.join(root, pat), recursive=True)
            if hits:
                return Path(sorted(hits)[-1])
    return None


# ------------------------------------------------------------------ styling

thin = Side(style="thin", color="D0D0D0")
HDR = PatternFill("solid", fgColor="5C3E34")
FLAG = PatternFill("solid", fgColor="FDE8E4")
ALT = PatternFill("solid", fgColor="FBF8EC")


def style(c, bold=False, color="000000", align="left", fmt=None, fill=None,
          size=10, wrap=False):
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_sheet(ws, title, subtitle, headers, rows, widths, fmts=None,
                flags=None):
    ws["A1"] = title
    ws["A1"].font = Font(name=ARIAL, size=14, bold=True, color="5C3E34")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=ARIAL, size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=False)

    r0 = 4
    for i, h in enumerate(headers, start=1):
        style(ws.cell(row=r0, column=i, value=h), bold=True, color="FFFFFF",
              align="center", fill=HDR, wrap=True)

    fmts = fmts or {}
    flags = flags or []
    for j, row in enumerate(rows):
        rr = r0 + 1 + j
        fill = ALT if j % 2 else None
        if j < len(flags) and flags[j]:
            fill = FLAG
        for i, v in enumerate(row, start=1):
            al = "right" if isinstance(v, (int, float)) else "left"
            style(ws.cell(row=rr, column=i, value=v), align=al,
                  fmt=fmts.get(i), fill=fill, wrap=(widths[i - 1] > 30))

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    ws.sheet_view.showGridLines = False
    if rows:
        ws.auto_filter.ref = (f"A{r0}:"
                              f"{get_column_letter(len(headers))}"
                              f"{r0 + len(rows)}")


def main() -> int:
    audit = find_audit()
    if audit is None:
        print("No Customer Discount Group Audit found.")
        return 1
    if not DB.exists():
        print(f"No database at {DB}")
        return 1

    print(f"audit: {audit}")
    a = pd.read_excel(audit, header=HEADER_ROW)
    a["ts"] = pd.to_datetime(a["Time"], errors="coerce")
    a = a.dropna(subset=["ts", "Customer ID", "Discount Description"])
    a["cid"] = a["Customer ID"].astype("Int64").astype(str)
    a["grp"] = a["Discount Description"].astype(str).str.strip()
    a["name"] = a["Customer Name"].astype(str).str.strip()

    # Open memberships only: last action for the pair is an Added.
    rows = []
    for (cid, grp), d in a.sort_values("ts").groupby(["cid", "grp"]):
        last = d.iloc[-1]
        if str(last.Action).strip() != "Added":
            continue
        rows.append({
            "cid": cid,
            "name": d["name"].iloc[-1],
            "group_name": grp,
            "group_kind": kind(grp),
            "added": last.ts,
            "by": str(last.get("Performed By", "")).strip(),
        })
    mem = pd.DataFrame(rows)
    print(f"open memberships: {len(mem):,}  customers: {mem.cid.nunique():,}")

    # --- transaction context ---------------------------------------------
    con = duckdb.connect(str(DB), read_only=True)
    con.register("mem_df", mem[["cid"]].drop_duplicates())
    tx = con.execute("""
        SELECT b.customer_key AS cid,
               COUNT(*)                                        AS baskets,
               COUNT(*) FILTER (WHERE COALESCE(b.discount_amt,0) > 0)
                                                               AS disc_baskets,
               SUM(COALESCE(b.discount_amt, 0))                AS discount,
               SUM(COALESCE(b.discount_amt,0)
                   - COALESCE(b.loyalty_redeem,0))             AS other_disc,
               SUM(b.basket_net)                               AS net,
               MODE(b.store_key)                               AS home_store,
               MAX(CAST(b.txn_ts AS DATE))                     AS last_seen
        FROM fact_basket b JOIN mem_df m ON b.customer_key = m.cid
        WHERE NOT b.is_return
        GROUP BY 1
    """).df()
    con.close()
    tx["home_store"] = tx.home_store.map(
        lambda k: STORE_LABEL.get(int(k), str(k)) if pd.notna(k) else "—")
    print(f"matched to transactions: {len(tx):,}")

    # --- per customer rollup ---------------------------------------------
    per = (mem.groupby("cid")
              .agg(name=("name", "last"),
                   groups=("group_name", lambda s: " | ".join(sorted(set(s)))),
                   kinds=("group_kind", lambda s: " | ".join(sorted(set(s)))),
                   n_groups=("group_name", "nunique"),
                   granted=("group_kind",
                            lambda s: sum(1 for x in set(s)
                                          if x != "Paid membership")),
                   granted_by=("by", lambda s: " | ".join(sorted(set(
                       x for x in s if x)))[:180]),
                   first_added=("added", "min"))
              .reset_index())
    per = per.merge(tx, on="cid", how="left")
    for c in ("baskets", "disc_baskets", "discount", "other_disc", "net"):
        per[c] = per[c].fillna(0)
    per["home_store"] = per.home_store.fillna("never transacted")
    per["depth"] = (per.other_disc
                    / (per.net + per.discount).replace(0, pd.NA) * 100)

    multi = per[per.n_groups > 1].sort_values(
        ["granted", "other_disc"], ascending=[False, False])
    gg = multi[multi.granted > 1]
    print(f"multi-group accounts: {len(multi):,}   granted+granted: {len(gg):,}")

    # --- concentrated use -------------------------------------------------
    grp = (mem.merge(tx, on="cid", how="left")
              .groupby(["group_name", "group_kind"])
              .agg(members=("cid", "nunique"),
                   disc_baskets=("disc_baskets", "sum"),
                   other_disc=("other_disc", "sum"))
              .reset_index())
    conc = grp[(grp.members <= 3) & (grp.disc_baskets >= 20)
               & (grp.group_kind != "Paid membership")]
    conc = conc.sort_values("other_disc", ascending=False)

    combos = (multi.groupby("kinds")
                   .agg(accounts=("cid", "nunique"),
                        other_disc=("other_disc", "sum"))
                   .reset_index()
                   .sort_values("accounts", ascending=False))

    # ------------------------------------------------------------ workbook
    wb = Workbook()
    today = datetime.date.today()
    src = f"Source: {audit.name} · generated {today:%d %b %Y} · internal use only"

    ws = wb.active
    ws.title = "Multi-group accounts"
    write_sheet(
        ws, "Accounts holding more than one discount entitlement",
        src + f" · {len(multi):,} accounts · rows shaded red hold 2+ GRANTED groups",
        ["Customer", "Customer ID", "Groups held", "Kinds", "# groups",
         "# granted", "Home store", "Discounted baskets", "Discount taken",
         "Non-loyalty portion", "Net spend", "Depth %", "Last seen",
         "Granted by"],
        [[r["name"], r.cid, r.groups, r.kinds, int(r.n_groups),
          int(r.granted), r.home_store, int(r.disc_baskets),
          round(float(r.discount), 2), round(float(r.other_disc), 2),
          round(float(r.net), 2),
          None if pd.isna(r.depth) else round(float(r.depth), 1),
          None if pd.isna(r.last_seen) else str(r.last_seen),
          r.granted_by]
         for _, r in multi.iterrows()],
        [24, 12, 46, 34, 9, 9, 15, 12, 13, 13, 13, 9, 12, 30],
        fmts={9: "$#,##0.00", 10: "$#,##0.00", 11: "$#,##0.00",
              12: "0.0"},
        flags=[bool(r.granted > 1) for _, r in multi.iterrows()])

    ws2 = wb.create_sheet("Granted + granted")
    write_sheet(
        ws2, "Two or more GRANTED entitlements — these should not exist",
        src + f" · {len(gg):,} accounts · a paid Frequent Flyer membership "
              "alongside one granted rate is excluded here",
        ["Customer", "Customer ID", "Groups held", "# granted", "Home store",
         "Discounted baskets", "Non-loyalty discount", "Net spend",
         "Granted by"],
        [[r["name"], r.cid, r.groups, int(r.granted), r.home_store,
          int(r.disc_baskets), round(float(r.other_disc), 2),
          round(float(r.net), 2), r.granted_by]
         for _, r in gg.iterrows()],
        [24, 12, 52, 10, 15, 12, 15, 13, 34],
        fmts={7: "$#,##0.00", 8: "$#,##0.00"})

    ws3 = wb.create_sheet("Concentrated use")
    write_sheet(
        ws3, "Groups with 3 or fewer members but heavy discount use",
        src + " · a rate intended for a business's staff being used as one "
              "person's standing discount",
        ["Group", "Kind", "Members", "Discounted baskets",
         "Non-loyalty discount"],
        [[r.group_name, r.group_kind, int(r.members), int(r.disc_baskets),
          round(float(r.other_disc), 2)] for _, r in conc.iterrows()],
        [40, 26, 10, 15, 16], fmts={5: "$#,##0.00"})

    ws4 = wb.create_sheet("Combinations")
    write_sheet(
        ws4, "Which entitlement pairings occur", src,
        ["Kinds held together", "Accounts", "Non-loyalty discount"],
        [[r.kinds, int(r.accounts), round(float(r.other_disc), 2)]
         for _, r in combos.iterrows()],
        [56, 11, 18], fmts={3: "$#,##0.00"})

    ws5 = wb.create_sheet("Notes")
    notes = [
        "HOW THIS WAS BUILT",
        "",
        f"Source audit: {audit.name}",
        f"Transaction data: {DB.name}",
        f"Generated: {today:%d %B %Y}",
        "",
        "Open memberships only. A customer x group pair counts only when the",
        "last recorded action for it is an Added. Anyone removed from a group",
        "is not flagged.",
        "",
        "Frequent Flyer is a $100 PURCHASED membership, not an earned tier, so",
        "it is labelled 'Paid membership'. Holding it alongside one granted",
        "rate is a policy question rather than an error, which is why the",
        "'Granted + granted' sheet excludes that combination.",
        "",
        "LIMITS OF THIS DATA",
        "",
        "Left-censored. The audit begins 2025-07-01. Anyone added to a group",
        "before that date has no Added event and will not appear here unless",
        "they were later removed and re-added. Roughly 75 memberships in the",
        "file begin with a Removed, which is direct evidence of this.",
        "",
        "Partial match. About 73% of audit customer IDs join to transaction",
        "records. The remainder never transacted, or transacted under a",
        "different customer key. Accounts showing 'never transacted' hold an",
        "entitlement but have no basket history.",
        "",
        "Home store is the store where the customer transacts most often, not",
        "a store assignment. Discount figures cover the customer's whole",
        "history, not only the period they held the entitlement.",
        "",
        "Non-loyalty discount is total discount minus the loyalty-offer",
        "portion. Loyalty is a SUBSET of total discount, so the two columns",
        "must never be added together.",
        "",
        "Whether entitlements STACK at the till is not established by this",
        "report. It shows who holds more than one, not that both were applied",
        "to the same basket.",
        "",
        "HANDLING",
        "",
        "This file names customers and the staff member who granted each",
        "entitlement. It is an internal document. It is deliberately not part",
        "of the Cerebral dashboard, which strips all customer identifiers",
        "before publishing.",
    ]
    for i, line in enumerate(notes, start=1):
        c = ws5.cell(row=i, column=1, value=line)
        c.font = Font(name=ARIAL, size=10,
                      bold=line.isupper() and bool(line.strip()),
                      color="5C3E34" if line.isupper() and line.strip()
                      else "333333")
        c.alignment = Alignment(vertical="top")
    ws5.column_dimensions["A"].width = 78
    ws5.sheet_view.showGridLines = False

    out = Path(f"Discount_Group_Audit_{today:%Y-%m-%d}.xlsx")
    wb.save(out)
    print(f"\nwrote {out}  ({out.stat().st_size/1024:.0f} KB)")
    print(f"  Multi-group accounts : {len(multi):,}")
    print(f"  Granted + granted    : {len(gg):,}")
    print(f"  Concentrated use     : {len(conc):,}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
