"""
vm_ingest.py — Visual Merchandising ingest for Cerebral.

Reads Visual_Merchandise_Data_2026.xlsx (one "Floor Set" sheet per store per
month) and unpivots every bay into a long placement table in tta.duckdb.

Tables written
--------------
fact_vm_placement   one row per (store, week, bay, position, brand) — physical
                    shelf placements; is_brand=false for accessory/signage rows
dim_vm_week         week_n -> week_start date per (store, month), so the tab
                    can join to POS weeks
dim_takeover        optional extra kiosk/featuring windows from takeovers.csv.
                    The canonical Takeover calendar is TAKEOVERS in
                    cerebral_public.py; publish_vm.py reads that directly and
                    unions this table on top (for one-offs not in the app).
vm_ingest_log       per-sheet parse report (rows parsed, warnings)

Usage
-----
    python vm_ingest.py                       # pulls the live Google Sheet (SHEET_ID)
                                              # to Visual_Merchandise_Data_2026.xlsx, then parses it
    python vm_ingest.py --no-download         # parse the local .xlsx as-is
    python vm_ingest.py --workbook path.xlsx --db ..\\tta.duckdb

Conventions this encodes (from the workbook layout)
---------------------------------------------------
* 14-column layout: A=Week, B=Bay, then three 4-col bands
  C-F = Top, G-J = Middle, K-N = Bottom, each = (Location, Brand, Swap, Impact)
* Shelf tier is taken from the COLUMN BAND, never from the label text.
* "Week N" in col A starts a week block; Bay in col B is forward-filled.
* Rows whose Bay ends in "- Note" are captured as notes, not placements.
* 5th Ave spotlight bay / six-window walls / SoHo kidney bean: position label
  is kept verbatim in `position_label`; tier still comes from the band.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path

import duckdb
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------

DEFAULT_WORKBOOK = Path("Visual_Merchandise_Data_2026.xlsx")
# The live workbook is a Google Sheet. Its xlsx export is pulled to
# DEFAULT_WORKBOOK before parsing (needs the sheet shared as "anyone with the
# link can view", or a signed-in browser download as a fallback).
SHEET_ID = "1PrYRWAYpsDLIEhP-To_SX-OExAHToa2WT9ZZlm9b3D8"
DEFAULT_DB = Path("..") / "tta.duckdb"
TAKEOVERS_CSV = Path("takeovers.csv")

# Week 1 = the first FLOOR_SET_WEEKDAY on or after the 1st of the month.
# 0=Mon ... 6=Sun. Monday keeps floor-set weeks aligned with the ISO weeks
# the rest of Cerebral uses (dash_brand_week is keyed on iso_year/iso_week).
FLOOR_SET_WEEKDAY = 0

# Per-(store, month) overrides for week_start when the deck didn't follow the
# rule. Optional file: store,month,week_n,week_start
WEEK_OVERRIDES_CSV = Path("vm_week_overrides.csv")

STORE_ALIASES = {
    "5th ave": "FIFTH", "fifth": "FIFTH",
    "soho": "SOHO",
    "usq": "USQ",
    "dtbk": "DTBK",
}
# Cerebral's STORES = {1: DTBK, 2: 5th Avenue, 3: Soho, 4: Union Square}
STORE_KEY = {"DTBK": 1, "FIFTH": 2, "SOHO": 3, "USQ": 4}

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}

BAND_TIER = {2: "Top", 6: "Middle", 10: "Bottom"}   # 0-based col of Location

# bay_raw (lower, stripped, "- Note" removed) -> bay_type
# Anything not matched falls through to "Other" and is logged.
BAY_TYPE_RULES: list[tuple[str, str]] = [
    (r"\bflower\b",                       "Flower"),
    (r"\bvape|concentrate\b",             "Vape"),
    (r"pre.?roll",                        "Pre-Roll"),
    (r"\bedib",                           "Edible"),
    (r"\bwellness\b",                     "Wellness"),
    (r"\baccessor|bong|glass wall\b",     "Accessory"),
    (r"kidney bean",                      "Kidney Bean"),
    (r"bubble",                           "Bubbles"),
    (r"\bcase\b",                         "Case"),
    (r"spotlight|roots of cannabis|future of flower|loyalty|display|podium|roundabout|grab and go|mother",
                                          "Spotlight"),
]


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def parse_sheet_name(name: str) -> tuple[str, int] | None:
    """'DTBK Floor Set August 2026' / 'SoHo September Floor Set 2026' ->
    ('DTBK', 8). Returns None for non-floor-set sheets."""
    low = name.lower()
    if "floor set" not in low:
        return None
    store = next((v for k, v in STORE_ALIASES.items() if low.startswith(k)), None)
    month = next((v for k, v in MONTHS.items() if k in low), None)
    if store is None or month is None:
        return None
    return store, month


def bay_type_for(bay_raw: str) -> str:
    low = bay_raw.lower()
    for pat, typ in BAY_TYPE_RULES:
        if re.search(pat, low):
            return typ
    return "Other"


_PAREN = re.compile(r"\(([^)]*)\)?")  # tolerates a missing close paren
_PTS = re.compile(r"\s*-?\s*([\d,]+)\s*pts\b", re.I)


def split_brand(cell: str) -> tuple[str, str]:
    """'Rythm (Flower)' -> ('Rythm', 'Flower');
    'Dogwalkers (Single) (Multipack)' -> ('Dogwalkers', 'Single; Multipack');
    '1906' stays '1906'."""
    hints = [h.strip() for h in _PAREN.findall(cell) if h.strip()]
    brand = _PAREN.sub("", cell)
    m = _PTS.search(brand)                      # 'Rythm - 1,000pts' (loyalty case)
    if m:
        hints.append(f"loyalty {m.group(1).replace(',', '')}pts")
        brand = _PTS.sub("", brand)
    return brand.strip(" -–|"), "; ".join(hints)


def brand_key(brand: str) -> str:
    """Join key: lowercase, alphanumerics only. 'B.Noble' == 'BNoble', 'PAX' == 'Pax'."""
    return re.sub(r"[^a-z0-9]", "", brand.lower())


_LIST_LIKE = re.compile(
    r",|^\d+\.\s|\"|\bposter\b|\bsignage\b|\bbackdrop\b|\bmenu\b|\bletter\b|"
    r"\bcard\b|\bjar\b|\bvase\b|\bpipes?\b|\bglass\b|\bcap\b|\bpen\b|\bbag\b|"
    r"\bgrinder\b|\bashtray\b|\btray\b|\bcase\b", re.I)
NON_BRAND_BAY_TYPES = {"Accessory"}


def looks_like_brand(cell: str, bay_type: str) -> bool:
    """False for accessory item lists, numbered SKU lists, signage, and
    obviously descriptive text — those stay in the table for coverage
    counts but must not be joined to POS brand sales."""
    if bay_type in NON_BRAND_BAY_TYPES:
        return False
    if _LIST_LIKE.search(cell):
        return False
    return len(cell) <= 40


def nth_weekday(year: int, month: int, weekday: int, n: int) -> dt.date:
    first = dt.date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + dt.timedelta(days=offset + 7 * (n - 1))


def load_week_overrides() -> dict[tuple[str, int, int], dt.date]:
    out = {}
    if WEEK_OVERRIDES_CSV.exists():
        with WEEK_OVERRIDES_CSV.open(newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                out[(r["store"].upper(), int(r["month"]), int(r["week_n"]))] = \
                    dt.date.fromisoformat(r["week_start"])
    return out


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------

@dataclass
class SheetResult:
    sheet: str
    store: str
    month: int
    placements: list[dict] = field(default_factory=list)
    notes: list[dict] = field(default_factory=list)
    weeks: set[int] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)


def parse_floor_set(ws, sheet: str, store: str, year: int, month: int) -> SheetResult:
    res = SheetResult(sheet, store, month)
    week_n: int | None = None
    bay_raw = ""
    bay_type = ""
    is_note_bay = False
    n_rows = 0

    for row in ws.iter_rows(min_row=2, max_col=14, values_only=True):
        cells = [norm(c) for c in row] + [""] * (14 - len(row))
        if not any(cells):
            continue
        n_rows += 1

        # Week marker (col A)
        m = re.match(r"week\s*(\d+)", cells[0], re.I)
        if m:
            week_n = int(m.group(1))
            res.weeks.add(week_n)
        elif cells[0] and week_n is None:
            # SoHo-Sept style shift: bay name landed in col A
            if not res.warnings:
                res.warnings.append("bay names in col A with no Week marker — looks like an unfilled template")

        # Bay (col B) forward-fill; also accept shifted layout where col A held it
        b = cells[1] or (cells[0] if not m and cells[0] else "")
        if b:
            bay_raw = b
            is_note_bay = bool(re.search(r"-\s*note$", bay_raw, re.I))
            bay_type = bay_type_for(re.sub(r"-\s*note$", "", bay_raw, flags=re.I))

        if week_n is None:
            continue  # nothing to attribute to yet

        if is_note_bay:
            text = " | ".join(c for c in cells[2:] if c)
            if text:
                res.notes.append(dict(store=store, month=month, week_n=week_n,
                                      bay_raw=bay_raw, note=text))
            continue

        # Three shelf bands
        for loc_col, tier in BAND_TIER.items():
            loc, brand_cell, swap_cell, impact = cells[loc_col:loc_col + 4]
            if not brand_cell and not swap_cell:
                continue
            base = dict(store=store, month=month, week_n=week_n,
                        bay_raw=bay_raw, bay_type=bay_type,
                        shelf_tier=tier, position_label=loc,
                        impact_note=impact)  # Impact col is left for Cerebral to compute
            # The "Swap" column is used as free-text commentary in practice
            # ("Pull requested", "Missing the 5pk"), so it rides along as a
            # comment rather than producing a second placement row.
            if brand_cell:
                brand, hints = split_brand(brand_cell)
                res.placements.append({**base, "brand_raw": brand_cell,
                                       "brand": brand, "brand_key": brand_key(brand),
                                       "product_hint": hints,
                                       "is_brand": looks_like_brand(brand_cell, bay_type),
                                       "comment": swap_cell})
            else:
                res.notes.append(dict(store=store, month=month, week_n=week_n,
                                      bay_raw=bay_raw, note=f"[{tier}/{loc}] {swap_cell}"))

    if n_rows and not res.weeks and not res.warnings:
        res.warnings.append("no 'Week N' markers found — sheet skipped for placements")
    if not n_rows:
        res.warnings.append("sheet is empty")
    return res


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def download_sheet(sheet_id: str, dest: Path) -> None:
    """Fetch the Google Sheet as xlsx. Google answers a login page (HTML) when
    the sheet is not link-shared — detect that rather than parse garbage."""
    import urllib.request
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print(f"Downloading sheet {sheet_id} ...", end="", flush=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            data = r.read()
    except Exception as e:  # HTTP 4xx, no network, proxy
        data = b""
        print(f" failed ({e})")
    if not data.startswith(b"PK"):
        raise SystemExit(
            "\nGoogle returned a web page instead of the workbook — the sheet is not "
            "shared as 'anyone with the link'. Either change sharing, or download it "
            f"from Sheets (File > Download > .xlsx) as {dest} and rerun with --no-download.")
    dest.write_bytes(data)
    print(f" {len(data)/1024:.0f} KB -> {dest}")


def run(workbook: Path, db: Path, year: int, download: bool = True) -> None:
    if download and SHEET_ID:
        download_sheet(SHEET_ID, workbook)
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook.resolve()}")
    wb = load_workbook(workbook, read_only=True, data_only=True)
    overrides = load_week_overrides()

    results: list[SheetResult] = []
    sheets = [(n, parse_sheet_name(n)) for n in wb.sheetnames]
    total = sum(1 for _, p in sheets if p)
    done = 0
    for name, parsed in sheets:
        if not parsed:
            continue
        store, month = parsed
        done += 1
        r = parse_floor_set(wb[name], name, store, year, month)
        results.append(r)
        print(f"[{done}/{total}] {name:<38} placements={len(r.placements):4d} "
              f"weeks={sorted(r.weeks)} {'⚠ ' + '; '.join(r.warnings) if r.warnings else ''}")

    # Week calendar
    week_rows = []
    for r in results:
        for n in sorted(r.weeks):
            start = overrides.get((r.store, r.month, n)) or \
                nth_weekday(year, r.month, FLOOR_SET_WEEKDAY, n)
            iso = start.isocalendar()
            week_rows.append(dict(store=r.store, store_key=STORE_KEY[r.store],
                                  month=r.month, week_n=n,
                                  week_start=start,
                                  week_end=start + dt.timedelta(days=6),
                                  iso_year=iso[0], iso_week=iso[1],
                                  source="override" if (r.store, r.month, n) in overrides else "rule"))

    placements = [p for r in results for p in r.placements]
    notes = [n for r in results for n in r.notes]
    log = [dict(sheet=r.sheet, store=r.store, month=r.month,
                placements=len(r.placements), notes=len(r.notes),
                weeks=",".join(map(str, sorted(r.weeks))),
                warnings=" | ".join(r.warnings),
                ingested_at=dt.datetime.now().isoformat(timespec="seconds"))
           for r in results]

    other = sorted({p["bay_raw"] for p in placements if p["bay_type"] == "Other"})
    if other:
        print(f"\n⚠ {len(other)} bay names fell through to bay_type='Other' — extend BAY_TYPE_RULES:")
        for b in other:
            print("   ", b)

    # Takeovers (stub: file may not exist yet)
    takeovers = []
    if TAKEOVERS_CSV.exists():
        with TAKEOVERS_CSV.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if not row.get("brand") or row["brand"].lstrip().startswith("#"):
                    continue  # blank or commented-out row
                takeovers.append(dict(
                    brand=row["brand"].strip(),
                    start_date=dt.date.fromisoformat(row["start_date"]),
                    end_date=dt.date.fromisoformat(row["end_date"]),
                    stores=row.get("stores", "").strip() or "USQ|DTBK|FIFTH|SOHO",
                    surface=row.get("surface", "").strip() or "kiosk",
                    source=row.get("source", "").strip() or "takeovers.csv",
                    airtable_record_id=row.get("airtable_record_id", "").strip(),
                    notes=row.get("notes", "").strip(),
                ))
    print(f"\nTakeovers loaded: {len(takeovers)} "
          f"({'from ' + str(TAKEOVERS_CSV) if TAKEOVERS_CSV.exists() else 'no takeovers.csv yet'})")

    # Write to DuckDB
    con = duckdb.connect(str(db))
    _replace(con, "fact_vm_placement", placements, """
        store VARCHAR, month INTEGER, week_n INTEGER,
        bay_raw VARCHAR, bay_type VARCHAR, shelf_tier VARCHAR, position_label VARCHAR,
        impact_note VARCHAR, brand_raw VARCHAR, brand VARCHAR, brand_key VARCHAR,
        product_hint VARCHAR, is_brand BOOLEAN, comment VARCHAR""")
    _replace(con, "vm_bay_note", notes, """
        store VARCHAR, month INTEGER, week_n INTEGER, bay_raw VARCHAR, note VARCHAR""")
    _replace(con, "dim_vm_week", week_rows, """
        store VARCHAR, store_key INTEGER, month INTEGER, week_n INTEGER,
        week_start DATE, week_end DATE, iso_year INTEGER, iso_week INTEGER,
        source VARCHAR""")
    _replace(con, "dim_takeover", takeovers, """
        brand VARCHAR, start_date DATE, end_date DATE, stores VARCHAR,
        surface VARCHAR, source VARCHAR, airtable_record_id VARCHAR, notes VARCHAR""")
    _replace(con, "vm_ingest_log", log, """
        sheet VARCHAR, store VARCHAR, month INTEGER, placements INTEGER, notes INTEGER,
        weeks VARCHAR, warnings VARCHAR, ingested_at VARCHAR""")
    con.close()
    print(f"\nWrote {len(placements)} placements, {len(notes)} notes, "
          f"{len(week_rows)} week rows, {len(takeovers)} takeovers -> {db}")


def _replace(con, table: str, rows: list[dict], schema: str) -> None:
    con.execute(f"CREATE OR REPLACE TABLE {table} ({schema})")
    if not rows:
        return
    cols = list(rows[0].keys())
    con.executemany(
        f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join('?' * len(cols))})",
        [[r[c] for c in cols] for r in rows])


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--no-download", action="store_true",
                    help="skip the Google Sheet export and parse --workbook as-is")
    a = ap.parse_args()
    run(a.workbook, a.db, a.year, download=not a.no_download)
